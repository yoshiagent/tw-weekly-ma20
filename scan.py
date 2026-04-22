"""
台股週K線過濾工具 v5
條件：
  1. 週 MA20 向上
  2. 過去 N 週內曾有大正乖離
  3. 目前股價已修正到 MA20 附近

圖表：週K線 + 布林通道（Bollinger Bands）+ 成交量
附加：三大法人週買賣超（大戶）+ 融資餘額週增減（散戶）+ 千張大戶比例（TDCC 集保分散表）
資料來源：TWSE + Yahoo Finance + TDCC 開放資料
"""

import yfinance as yf
import pandas as pd
import numpy as np
import requests
import sys
import os
import json
import time
from datetime import datetime, timedelta

# ── 參數設定 ──────────────────────────────────────────────────────────────────
LOOKBACK_WEEKS    = 16    # 往回看幾週判斷「曾有大正乖離」
BIG_DEV_THRESHOLD = 0.15  # 大正乖離門檻 (15%)
NEAR_MA20_MAX     = 0.05  # 目前乖離上限 +5%
NEAR_MA20_MIN     = -0.03 # 目前乖離下限 -3%
MA20_SLOPE_WEEKS  = 4     # MA20 斜率計算週數
CHART_WEEKS       = 60    # 圖表顯示週數
BB_STD            = 2     # 布林通道標準差倍數

# ── 1. 取得台股上市清單 ───────────────────────────────────────────────────────
def get_twse_stock_list():
    print("取得台股上市清單...", flush=True)
    url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
    try:
        r = requests.get(url, timeout=30)
        data = r.json()
        stocks = []
        for item in data:
            code = item.get("Code", "")
            name = item.get("Name", "")
            if code.isdigit() and len(code) == 4:
                stocks.append({"code": code, "name": name})
        print(f"   共 {len(stocks)} 檔上市股票", flush=True)
        return stocks
    except Exception as e:
        print(f"   無法取得清單：{e}", flush=True)
        return []

# ── 2. 計算布林通道 ───────────────────────────────────────────────────────────
def calc_bollinger(close_series, window=20, std_mult=2):
    ma   = close_series.rolling(window).mean()
    std  = close_series.rolling(window).std(ddof=0)
    upper = ma + std_mult * std
    lower = ma - std_mult * std
    return ma, upper, lower

# ── 3. 計算特定價格區間的加權成交量（Volume Profile 近似）────────────────────
def calc_zone_volume(highs, lows, volumes, z_low, z_high):
    """
    假設每根 K 棒成交量均勻分布在 [low, high] 區間，
    估算落在 [z_low, z_high] 的成交量占比。
    輸入為 numpy array；回傳整數（股）。
    """
    if z_high <= z_low:
        return 0
    span    = highs - lows                                   # 每根 K 棒的價格跨幅
    ol_low  = np.maximum(z_low,  lows)
    ol_high = np.minimum(z_high, highs)
    overlap = np.maximum(0.0, ol_high - ol_low)
    valid   = span > 0
    prop    = np.where(valid, overlap / np.where(valid, span, 1.0), 0.0)
    return int(np.sum(volumes * prop))

# ── 4. 批次下載週K線並套用篩選條件 ───────────────────────────────────────────
def scan_stocks(stocks, batch_size=150):
    results = []
    total = len(stocks)

    for i in range(0, total, batch_size):
        batch = stocks[i : i + batch_size]
        tickers = [f"{s['code']}.TW" for s in batch]
        name_map = {f"{s['code']}.TW": s["name"] for s in batch}

        pct = (i + len(batch)) / total * 100
        print(f"   下載中 {i+1}~{i+len(batch)}/{total}（{pct:.0f}%）...", flush=True)

        try:
            raw = yf.download(
                tickers,
                period="2y",
                interval="1wk",
                auto_adjust=True,
                progress=False,
                group_by="ticker",
                threads=True,
            )
        except Exception as e:
            print(f"   批次下載失敗：{e}", flush=True)
            continue

        for ticker in tickers:
            try:
                single = len(tickers) == 1
                close  = raw["Close"]          if single else raw[ticker]["Close"]
                open_  = raw["Open"]           if single else raw[ticker]["Open"]
                high   = raw["High"]           if single else raw[ticker]["High"]
                low    = raw["Low"]            if single else raw[ticker]["Low"]
                volume = raw["Volume"]         if single else raw[ticker]["Volume"]

                close  = close.dropna()
                if len(close) < 25:
                    continue

                # 篩選邏輯
                ma20 = close.rolling(20).mean()
                ma20_clean = ma20.dropna()
                if len(ma20_clean) < MA20_SLOPE_WEEKS + 2:
                    continue

                last_close = float(close.iloc[-1])
                last_ma20  = float(ma20_clean.iloc[-1])
                old_ma20   = float(ma20_clean.iloc[-1 - MA20_SLOPE_WEEKS])

                if not (last_ma20 > old_ma20):
                    continue

                recent_close = close.iloc[-LOOKBACK_WEEKS:]
                recent_ma20  = ma20.iloc[-LOOKBACK_WEEKS:]
                dev_series   = (recent_close - recent_ma20) / recent_ma20
                max_dev = float(dev_series.max())
                if max_dev < BIG_DEV_THRESHOLD:
                    continue

                current_dev = (last_close - last_ma20) / last_ma20
                if not (NEAR_MA20_MIN <= current_dev <= NEAR_MA20_MAX):
                    continue

                # 條件 4：收盤價 >= 50
                if last_close < 50:
                    continue

                ma20_slope_pct   = (last_ma20 - old_ma20) / old_ma20 * 100
                peak_idx         = int(dev_series.argmax())
                weeks_since_peak = LOOKBACK_WEEKS - 1 - peak_idx

                # ── 壓力量 / 支撐量（Volume Profile 近似）────────────────────
                # 前高 = LOOKBACK_WEEKS 內的最高 high
                prev_high_price = float(high.iloc[-LOOKBACK_WEEKS:].max())
                D_val = prev_high_price - last_close

                # 使用全部可用週K資料進行估算
                _df_vp = pd.DataFrame({
                    "high":   high,
                    "low":    low,
                    "volume": volume,
                }).dropna()
                _h = _df_vp["high"].astype(float).values
                _l = _df_vp["low"].astype(float).values
                _v = _df_vp["volume"].astype(float).values

                if D_val > 0 and len(_df_vp) > 0:
                    resist_vol = calc_zone_volume(_h, _l, _v,
                                                  last_close, prev_high_price)
                    support_vol = calc_zone_volume(_h, _l, _v,
                                                   last_close - D_val, last_close)
                    vol_ratio = round(support_vol / resist_vol, 2) if resist_vol > 0 else None
                else:
                    resist_vol  = 0
                    support_vol = 0
                    vol_ratio   = None

                # 計算布林通道
                _, bb_upper, bb_lower = calc_bollinger(close, window=20, std_mult=BB_STD)

                # 取最近 CHART_WEEKS 週的 OHLCV + 指標，對齊索引
                df = pd.DataFrame({
                    "open":     open_,
                    "high":     high,
                    "low":      low,
                    "close":    close,
                    "volume":   volume,
                    "ma20":     ma20,
                    "bb_upper": bb_upper,
                    "bb_lower": bb_lower,
                }).dropna(subset=["close"]).tail(CHART_WEEKS)

                candles, vol_bars, ma_line, bbu_line, bbl_line = [], [], [], [], []
                for dt, row in df.iterrows():
                    t = dt.strftime("%Y-%m-%d")
                    candles.append({
                        "time":  t,
                        "open":  round(float(row["open"]),  2),
                        "high":  round(float(row["high"]),  2),
                        "low":   round(float(row["low"]),   2),
                        "close": round(float(row["close"]), 2),
                    })
                    vol_bars.append({
                        "time":  t,
                        "value": int(row["volume"]) if not pd.isna(row["volume"]) else 0,
                        "color": "#f87171" if float(row["close"]) >= float(row["open"]) else "#34d399",
                    })
                    if not pd.isna(row["ma20"]):
                        ma_line.append( {"time": t, "value": round(float(row["ma20"]),  2)})
                        bbu_line.append({"time": t, "value": round(float(row["bb_upper"]), 2)})
                        bbl_line.append({"time": t, "value": round(float(row["bb_lower"]), 2)})

                code = ticker.replace(".TW", "")
                results.append({
                    "code":             code,
                    "name":             name_map.get(ticker, ""),
                    "close":            round(last_close, 2),
                    "ma20":             round(last_ma20,  2),
                    "current_dev":      round(current_dev * 100, 2),
                    "peak_dev":         round(max_dev * 100,     2),
                    "weeks_since_peak": weeks_since_peak,
                    "ma20_slope_pct":   round(ma20_slope_pct,    2),
                    "prev_high":        round(prev_high_price,   2),
                    "resist_vol":       resist_vol,    # 單位：股（shares）
                    "support_vol":      support_vol,
                    "vol_ratio":        vol_ratio,
                    "chart": {
                        "candles":  candles,
                        "volume":   vol_bars,
                        "ma20":     ma_line,
                        "bb_upper": bbu_line,
                        "bb_lower": bbl_line,
                    },
                })

            except Exception:
                continue

    return results

# ── 5. 取得近期有資料的交易日清單 ─────────────────────────────────────────────
def get_trading_days(n=10):
    """找最近 n 個有三大法人資料的交易日"""
    headers = {"User-Agent": "Mozilla/5.0"}
    days = []
    d = datetime.now() - timedelta(days=1)
    while len(days) < n:
        if d.weekday() < 5:
            date_str = d.strftime("%Y%m%d")
            try:
                r = requests.get(
                    "https://www.twse.com.tw/rwd/zh/fund/T86",
                    params={"response": "json", "date": date_str, "selectType": "ALL"},
                    headers=headers, timeout=10
                )
                if r.json().get("total", 0) > 0:
                    days.append(date_str)
            except Exception:
                pass
        d -= timedelta(days=1)
    return days  # 由新到舊

# ── 6. 三大法人週買賣超（大戶）─────────────────────────────────────────────────
def fetch_institutional_weekly(trading_days):
    """
    取最近兩週共 10 個交易日的 T86 資料
    回傳 dict: code -> {"this_week": 本週合計萬股, "prev_week": 上週合計萬股}
    """
    headers = {"User-Agent": "Mozilla/5.0"}
    this_week_days = trading_days[:5]   # 最新 5 天
    prev_week_days = trading_days[5:10] # 前一週 5 天

    def sum_week(days):
        totals = {}
        for date_str in days:
            try:
                r = requests.get(
                    "https://www.twse.com.tw/rwd/zh/fund/T86",
                    params={"response": "json", "date": date_str, "selectType": "ALL"},
                    headers=headers, timeout=10
                )
                data = r.json()
                for row in data.get("data", []):
                    code = row[0]
                    try:
                        net = int(row[18].replace(",", "").replace(" ", ""))
                        totals[code] = totals.get(code, 0) + net
                    except Exception:
                        pass
                time.sleep(0.3)
            except Exception:
                pass
        return totals

    print("   抓三大法人本週資料…", flush=True)
    this_w = sum_week(this_week_days)
    print("   抓三大法人上週資料…", flush=True)
    prev_w = sum_week(prev_week_days)

    result = {}
    all_codes = set(this_w) | set(prev_w)
    for code in all_codes:
        result[code] = {
            "this_week": round(this_w.get(code, 0) / 10000, 1),   # 萬股
            "prev_week": round(prev_w.get(code, 0) / 10000, 1),
        }
    return result

# ── 7. 融資餘額週增減（散戶代理指標）─────────────────────────────────────────
MARGIN_CACHE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "margin_cache.json")
TDCC_CACHE_PATH   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tdcc_cache.json")

def fetch_margin_current():
    """從 TWSE openapi 取得當日融資餘額（含前日）"""
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        r = requests.get(
            "https://openapi.twse.com.tw/v1/exchangeReport/MI_MARGN",
            headers=headers, timeout=15
        )
        data = r.json()
        result = {}
        for item in data:
            code = item.get("股票代號", "")
            if code.isdigit() and len(code) == 4:
                try:
                    result[code] = int(item.get("融資今日餘額", "0").replace(",", "") or "0")
                except Exception:
                    result[code] = 0
        return result
    except Exception as e:
        print(f"   ⚠ 融資資料抓取失敗：{e}", flush=True)
        return {}

def get_margin_weekly_change():
    """
    比對本次與上次快取的融資餘額
    回傳 dict: code -> 週增減（張，= 本週餘額 - 上週餘額）
    """
    # 讀上次快取
    prev_margin = {}
    if os.path.exists(MARGIN_CACHE_PATH):
        try:
            with open(MARGIN_CACHE_PATH, "r", encoding="utf-8") as f:
                prev_margin = json.load(f)
        except Exception:
            pass

    print("   抓融資餘額資料…", flush=True)
    curr_margin = fetch_margin_current()

    # 計算週增減（單位：張）
    weekly_chg = {}
    if prev_margin:
        for code in curr_margin:
            curr = curr_margin.get(code, 0)
            prev = prev_margin.get(code, curr)  # 若上週無資料則視為 0 增減
            weekly_chg[code] = curr - prev
    else:
        # 第一次執行，沒有歷史資料
        for code in curr_margin:
            weekly_chg[code] = None  # None = 無法計算

    # 儲存本次快取（供下週比對）
    try:
        with open(MARGIN_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(curr_margin, f, ensure_ascii=False)
    except Exception as e:
        print(f"   ⚠ 快取儲存失敗：{e}", flush=True)

    return weekly_chg

# ── 8. TDCC 集保分散表 — 千張大戶比例 ────────────────────────────────────────
def fetch_tdcc_big_holder():
    """
    從 TDCC 開放資料下載集保分散表 CSV
    Level 15 = 持股 1,000,001 股以上（千張以上）
    回傳 dict: code -> {"pct": float, "date": str}
    """
    url = "https://opendata.tdcc.com.tw/getOD.ashx?id=1-5"
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        r = requests.get(url, timeout=60, headers=headers)
        text = r.content.decode("utf-8-sig")
        lines = text.splitlines()

        result = {}
        data_date = None

        for line in lines[1:]:   # 跳過標題列
            parts = line.split(",")
            if len(parts) < 6:
                continue
            date_str = parts[0].strip()
            code     = parts[1].strip()   # 去除尾端空格
            level    = parts[2].strip()

            if not data_date and date_str:
                data_date = date_str

            # 只保留 4 碼純數字（一般上市股票）
            if not (code.isdigit() and len(code) == 4):
                continue

            # Level 15 = 千張以上（1,000,001 股以上）
            if level == "15":
                try:
                    pct = float(parts[5].strip())
                    result[code] = {"pct": pct, "date": date_str}
                except Exception:
                    pass

        print(f"   TDCC 集保資料日期：{data_date}，共 {len(result)} 檔", flush=True)
        return result
    except Exception as e:
        print(f"   ⚠ TDCC 資料抓取失敗：{e}", flush=True)
        return {}

def get_tdcc_weekly_change():
    """
    比對本次與上次快取的千張大戶比例，計算週增減
    回傳 dict: code -> {"pct": float, "chg": float or None}
    """
    prev_tdcc = {}
    if os.path.exists(TDCC_CACHE_PATH):
        try:
            with open(TDCC_CACHE_PATH, "r", encoding="utf-8") as f:
                prev_tdcc = json.load(f)
        except Exception:
            pass

    print("   抓 TDCC 千張大戶比例…", flush=True)
    curr_tdcc = fetch_tdcc_big_holder()

    result = {}
    for code, curr_data in curr_tdcc.items():
        pct = curr_data["pct"]
        if prev_tdcc and code in prev_tdcc:
            prev_pct = prev_tdcc[code].get("pct", pct)
            chg = round(pct - prev_pct, 2)
        else:
            chg = None   # 第一次執行
        result[code] = {"pct": pct, "chg": chg}

    # 儲存本次快取（供下週比對）
    try:
        with open(TDCC_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(curr_tdcc, f, ensure_ascii=False)
    except Exception as e:
        print(f"   ⚠ TDCC 快取儲存失敗：{e}", flush=True)

    return result

# ── 9. 產生 HTML ──────────────────────────────────────────────────────────────
def gen_html(results, scan_time):
    results.sort(key=lambda x: abs(x["current_dev"]))

    # 把圖表資料分離，只把 code→chart 的映射嵌入 JS
    chart_data_map = {r["code"]: r["chart"] for r in results}
    chart_data_json = json.dumps(chart_data_map, ensure_ascii=False)

    rows = ""
    for r in results:
        dev_color = "pos-green" if r["current_dev"] >= 0 else "neg-red"
        dev_str   = f"{r['current_dev']:+.2f}%"
        peak_str  = f"{r['peak_dev']:+.2f}%"
        slope_str = f"{'▲' if r['ma20_slope_pct'] > 0 else '▼'} {r['ma20_slope_pct']:.2f}%"

        # 三大法人週買賣超（大戶）
        inst = r.get("institutional") or {}
        tw = inst.get("this_week")
        if tw is not None:
            inst_color = "pos-red" if tw > 0 else ("neg-green" if tw < 0 else "neutral")
            inst_str   = f"{tw:+,.1f}"
        else:
            inst_color, inst_str = "neutral", "—"

        # 融資週增減（散戶）
        mc = r.get("margin_chg")
        if mc is None:
            margin_color, margin_str = "neutral", "初次"
        else:
            margin_color = "neg-green" if mc > 0 else ("pos-red" if mc < 0 else "neutral")
            margin_str   = f"{mc:+,}"

        # 千張大戶比例 & 週增減（TDCC）
        tdcc = r.get("tdcc") or {}
        big_pct = tdcc.get("pct")
        big_chg = tdcc.get("chg")
        if big_pct is None:
            big_pct_str = "—"
            big_pct_color = "neutral"
        else:
            big_pct_str   = f"{big_pct:.2f}%"
            big_pct_color = "neutral"
        if big_chg is None:
            big_chg_str = "初次"
            big_chg_color = "neutral"
        elif big_chg > 0:
            big_chg_str   = f"+{big_chg:.2f}%"
            big_chg_color = "pos-red"    # 大戶增加 → 紅（看漲）
        elif big_chg < 0:
            big_chg_str   = f"{big_chg:.2f}%"
            big_chg_color = "neg-green"  # 大戶減少 → 綠（偏弱）
        else:
            big_chg_str   = "0.00%"
            big_chg_color = "neutral"

        # ── 壓力量 / 支撐量 / 比值 ────────────────────────────────────────
        # 顯示單位：萬張（1萬張 = 10,000,000 股）
        def _vol_fmt(v):
            if v is None or v == 0:
                return "—"
            wan = v / 1e7
            if wan >= 100:
                return f"{wan:.0f}"
            elif wan >= 10:
                return f"{wan:.1f}"
            else:
                return f"{wan:.2f}"

        rv = r.get("resist_vol",  0)
        sv = r.get("support_vol", 0)
        vr = r.get("vol_ratio")

        resist_str  = _vol_fmt(rv)
        support_str = _vol_fmt(sv)

        if vr is None:
            ratio_str   = "—"
            ratio_color = "neutral"
        elif vr >= 1.5:
            ratio_str   = f"{vr:.2f}"
            ratio_color = "ratio-high"    # 支撐強 → 深綠
        elif vr >= 1.0:
            ratio_str   = f"{vr:.2f}"
            ratio_color = "neg-green"     # 支撐 > 壓力 → 綠
        elif vr >= 0.67:
            ratio_str   = f"{vr:.2f}"
            ratio_color = "neutral"       # 接近平衡
        else:
            ratio_str   = f"{vr:.2f}"
            ratio_color = "neg-red"       # 壓力明顯 → 紅

        rows += f"""
        <tr data-code="{r['code']}" data-name="{r['name']}">
          <td class="code">{r['code']}</td>
          <td>{r['name']}</td>
          <td class="num">{r['close']}</td>
          <td class="num">{r['ma20']}</td>
          <td class="num {dev_color}">{dev_str}</td>
          <td class="num peak">{peak_str}</td>
          <td class="num">{r['weeks_since_peak']} 週前</td>
          <td class="num up">{slope_str}</td>
          <td class="num {inst_color}">{inst_str}</td>
          <td class="num {margin_color}">{margin_str}</td>
          <td class="num {big_pct_color}">{big_pct_str}</td>
          <td class="num {big_chg_color}">{big_chg_str}</td>
          <td class="num neutral">{resist_str}</td>
          <td class="num neutral">{support_str}</td>
          <td class="num {ratio_color}">{ratio_str}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>台股週K MA20 回測清單｜正乖離修正</title>
  <script src="https://unpkg.com/lightweight-charts@4.2.0/dist/lightweight-charts.standalone.production.js"></script>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      background: #080e18;
      color: #e2e8f0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      line-height: 1.7;
      padding: 32px 20px;
    }}
    h1 {{ font-size: 1.6rem; color: #3b82f6; margin-bottom: 6px; }}
    .subtitle {{ color: #94a3b8; font-size: 0.9rem; margin-bottom: 24px; }}

    .stats {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 28px; }}
    .stat-card {{
      background: #131f2e; border: 1px solid #2d3f55;
      border-radius: 10px; padding: 14px 22px; min-width: 140px;
    }}
    .stat-label {{ font-size: 0.78rem; color: #94a3b8; margin-bottom: 4px; }}
    .stat-value {{ font-size: 1.5rem; font-weight: 700; color: #3b82f6; }}

    .callout {{
      background: #0f2040; border: 1px solid rgba(59,130,246,0.3);
      border-radius: 8px; color: #93c5fd; font-size: 0.85rem;
      padding: 14px 18px; margin-bottom: 24px; line-height: 1.8;
    }}
    .callout strong {{ color: #60a5fa; }}

    .params {{ display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 24px; }}
    .param-tag {{
      background: #0f2040; border: 1px solid rgba(59,130,246,0.25);
      border-radius: 20px; color: #7dd3fc; font-size: 0.78rem; padding: 4px 12px;
    }}

    .search-bar {{ margin-bottom: 16px; }}
    .search-bar input {{
      background: #131f2e; border: 1px solid #2d3f55; border-radius: 6px;
      color: #e2e8f0; font-size: 0.9rem; padding: 8px 14px; width: 280px; outline: none;
    }}
    .search-bar input:focus {{ border-color: #3b82f6; }}

    .table-wrap {{ overflow-x: auto; border-radius: 10px; border: 1px solid #2d3f55; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 0.9rem; }}
    thead tr {{ background: #0f2040; }}
    th {{
      padding: 12px 16px; text-align: left; color: #93c5fd;
      font-weight: 600; font-size: 0.82rem; white-space: nowrap;
      cursor: pointer; user-select: none;
    }}
    th:hover {{ color: #3b82f6; }}
    tbody tr {{
      border-top: 1px solid #1a2d42; transition: background 0.15s; cursor: pointer;
    }}
    tbody tr:hover {{ background: #162030; }}
    tbody tr.active {{ background: #0f2040; outline: 1px solid #3b82f6; }}
    td {{ padding: 10px 16px; white-space: nowrap; }}
    .code {{ color: #fb923c; font-weight: 600; font-family: monospace; font-size: 0.95rem; }}
    .num       {{ text-align: right; }}
    .pos-green  {{ color: #34d399; }}
    .neg-red    {{ color: #f87171; }}
    .pos-red    {{ color: #f87171; }}
    .neg-green  {{ color: #34d399; }}
    .neutral    {{ color: #94a3b8; }}
    .peak       {{ color: #fb923c; }}
    .up         {{ color: #3b82f6; }}
    .ratio-high {{ color: #4ade80; font-weight: 700; }}

    /* ── 圖表 Modal ────────────────────────────────────── */
    .modal-overlay {{
      display: none;
      position: fixed; inset: 0;
      background: rgba(0,0,0,0.75);
      z-index: 1000;
      align-items: center;
      justify-content: center;
      padding: 16px;
    }}
    .modal-overlay.open {{ display: flex; }}

    .modal {{
      background: #0d1826;
      border: 1px solid #2d3f55;
      border-radius: 14px;
      width: 100%;
      max-width: 900px;
      max-height: 90vh;
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }}

    .modal-header {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      padding: 14px 20px;
      border-bottom: 1px solid #2d3f55;
      gap: 12px;
      flex-shrink: 0;
    }}
    .modal-title {{
      font-size: 1.05rem;
      font-weight: 600;
      color: #e2e8f0;
    }}
    .modal-title span {{ color: #fb923c; margin-right: 8px; font-family: monospace; }}

    .modal-legend {{
      display: flex; gap: 16px; flex-wrap: wrap; font-size: 0.78rem;
    }}
    .legend-item {{ display: flex; align-items: center; gap: 5px; color: #94a3b8; }}
    .legend-dot {{
      width: 20px; height: 3px; border-radius: 2px; flex-shrink: 0;
    }}

    .modal-close {{
      background: none; border: none; color: #94a3b8;
      font-size: 1.4rem; cursor: pointer; line-height: 1;
      padding: 2px 6px; border-radius: 4px;
      flex-shrink: 0;
    }}
    .modal-close:hover {{ color: #e2e8f0; background: #1e2d40; }}

    .chart-area {{
      padding: 0 4px 4px;
      flex: 1;
      min-height: 0;
    }}

    #main-chart  {{ width: 100%; height: 340px; }}
    #vol-chart   {{ width: 100%; height: 100px; margin-top: 2px; }}

    .modal-footer {{
      padding: 8px 20px;
      border-top: 1px solid #1a2d42;
      font-size: 0.75rem;
      color: #4a6080;
      display: flex;
      justify-content: space-between;
      flex-shrink: 0;
    }}
    .modal-footer a {{ color: #3b82f6; text-decoration: none; }}
    .modal-footer a:hover {{ text-decoration: underline; }}

    footer {{
      margin-top: 36px; padding-top: 16px;
      border-top: 1px solid #2d3f55; font-size: 0.78rem; color: #94a3b8;
    }}
  </style>
</head>
<body>

<h1>台股週K MA20 ｜正乖離修正回測清單</h1>
<p class="subtitle">資料來源：Yahoo Finance 週K線（還原權值）　｜　點擊列查看週K圖表</p>

<div class="stats">
  <div class="stat-card" style="border-color: rgba(59,130,246,0.5);">
    <div class="stat-label">最後更新</div>
    <div class="stat-value" style="font-size:0.95rem; color:#60a5fa; padding-top:4px; letter-spacing:0.02em;">{scan_time}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">符合條件股票</div>
    <div class="stat-value">{len(results)}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">大正乖離門檻</div>
    <div class="stat-value" style="color:#fb923c">≥ {int(BIG_DEV_THRESHOLD*100)}%</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">目前乖離範圍</div>
    <div class="stat-value" style="font-size:1rem; padding-top:6px; color:#34d399">
      {int(NEAR_MA20_MIN*100)}% ～ +{int(NEAR_MA20_MAX*100)}%
    </div>
  </div>
  <div class="stat-card">
    <div class="stat-label">回看週期</div>
    <div class="stat-value" style="color:#7dd3fc">{LOOKBACK_WEEKS} 週</div>
  </div>
</div>

<div class="callout">
  <strong>篩選邏輯說明</strong><br>
  ① <strong>週 MA20 向上（Moving Average 20-week）</strong>：當前 MA20 高於 {MA20_SLOPE_WEEKS} 週前的 MA20，確認均線處於上升趨勢<br>
  ② <strong>曾有大正乖離（Positive Deviation）</strong>：過去 {LOOKBACK_WEEKS} 週內，收盤價曾超越 MA20 達 <strong>+{int(BIG_DEV_THRESHOLD*100)}% 以上</strong><br>
  ③ <strong>修正到 MA20 附近（Pullback to MA）</strong>：目前乖離率介於 <strong>{int(NEAR_MA20_MIN*100)}% 至 +{int(NEAR_MA20_MAX*100)}%</strong><br>
  📊 點擊任一列可查看 <strong>週K線圖 + 布林通道（Bollinger Bands, {BB_STD}σ）</strong><br>
  👥 <strong>法人(萬股)</strong>：三大法人（外資＋投信＋自營）本週合計買賣超，正值紅色＝買超，負值綠色＝賣超<br>
  💳 <strong>融資週增減(張)</strong>：融資餘額與上週比較，正值紅色＝融資增加（散戶積極做多），負值綠色＝融資減少（健康去槓桿）<br>
  🏦 <strong>千張大戶比</strong>：集保分散表 Level-15（持股千張以上）占全部庫存比例，反映法人/大戶集中度；<strong>週增減</strong>正值紅色＝大戶持續加碼，負值綠色＝大戶減倉（資料來源：TDCC 開放資料）<br>
  📊 <strong>壓力量／支撐量</strong>：以週K線均勻分布法估算，D＝前高－收盤；壓力量＝[收盤, 前高]區間歷史成交量；支撐量＝[收盤－D, 收盤]等寬區間歷史成交量；<strong>支/壓比值</strong>＞1.5 深綠＝支撐顯著強，＞1.0 綠＝支撐佔優，＜0.67 紅＝壓力明顯
</div>

<div class="params">
  <span class="param-tag">回看 {LOOKBACK_WEEKS} 週</span>
  <span class="param-tag">大乖離門檻 +{int(BIG_DEV_THRESHOLD*100)}%</span>
  <span class="param-tag">MA20 附近 {int(NEAR_MA20_MIN*100)}%～+{int(NEAR_MA20_MAX*100)}%</span>
  <span class="param-tag">布林通道 MA20 ± {BB_STD}σ</span>
  <span class="param-tag">圖表顯示 {CHART_WEEKS} 週</span>
</div>

<div class="search-bar">
  <input type="text" id="search" placeholder="搜尋代碼或名稱..." oninput="filterTable()">
</div>

<div class="table-wrap">
  <table id="mainTable">
    <thead>
      <tr>
        <th onclick="sortTable(0)">代碼 ↕</th>
        <th onclick="sortTable(1)">名稱 ↕</th>
        <th onclick="sortTable(2)">收盤 ↕</th>
        <th onclick="sortTable(3)">MA20 ↕</th>
        <th onclick="sortTable(4)" title="目前收盤與週MA20的乖離率">乖離 ↕</th>
        <th onclick="sortTable(5)" title="回看期間內最大正乖離">峰乖離 ↕</th>
        <th onclick="sortTable(6)" title="距離峰值已過幾週">距峰 ↕</th>
        <th onclick="sortTable(7)" title="MA20與4週前相比的斜率">斜率 ↕</th>
        <th onclick="sortTable(8)" title="三大法人本週買賣超合計（萬股）紅=買超 綠=賣超">法人 ↕</th>
        <th onclick="sortTable(9)" title="融資餘額週增減（張）綠=減少 紅=增加">融資 ↕</th>
        <th onclick="sortTable(10)" title="集保分散表千張以上持股比例（TDCC Level-15）">千張比 ↕</th>
        <th onclick="sortTable(11)" title="千張大戶比週增減，正值紅=加碼，負值綠=減倉">千張變 ↕</th>
        <th onclick="sortTable(12)" title="前高至收盤區間的加權成交量（萬張），代表上方套牢壓力">壓力 ↕</th>
        <th onclick="sortTable(13)" title="收盤至(收盤-D)等寬區間的加權成交量（萬張），代表下方支撐籌碼">支撐 ↕</th>
        <th onclick="sortTable(14)" title="支撐量÷壓力量，>1.5深綠=支撐強，>1.0綠=支撐佔優，<0.67紅=壓力明顯">支/壓 ↕</th>
      </tr>
    </thead>
    <tbody>
      {rows}
    </tbody>
  </table>
</div>

<!-- ── 圖表 Modal ────────────────────────────────────────────────────────── -->
<div class="modal-overlay" id="modalOverlay" onclick="closeModalOnBg(event)">
  <div class="modal">
    <div class="modal-header">
      <div class="modal-title" id="modalTitle"></div>
      <div class="modal-legend">
        <div class="legend-item"><div class="legend-dot" style="background:#3b82f6"></div>MA20</div>
        <div class="legend-item"><div class="legend-dot" style="background:#fb923c; opacity:.8"></div>BB 上軌</div>
        <div class="legend-item"><div class="legend-dot" style="background:#a78bfa; opacity:.8"></div>BB 下軌</div>
      </div>
      <button class="modal-close" onclick="closeModal()">✕</button>
    </div>
    <div class="chart-area">
      <div id="main-chart"></div>
      <div id="vol-chart"></div>
    </div>
    <div class="modal-footer">
      <span>週K線（還原權值）｜布林通道 MA20 ± {BB_STD}σ</span>
      <a id="yahooLink" href="#" target="_blank" rel="noopener">在 Yahoo Finance 查看 ↗</a>
    </div>
  </div>
</div>

<footer>
  資料來源：<strong>Yahoo Finance</strong> 週K線（auto_adjust 還原權值）、台灣證券交易所（TWSE）上市股票清單<br>
  圖表：<strong>TradingView Lightweight Charts</strong>｜注意：資料僅供參考，不構成任何投資建議。
</footer>

<script>
// ── 圖表資料 ─────────────────────────────────────────────────────────────────
const CHART_DATA = {chart_data_json};

// ── 排序 ──────────────────────────────────────────────────────────────────────
let sortDir = {{}};
function sortTable(col) {{
  const tb   = document.querySelector("#mainTable tbody");
  const rows = Array.from(tb.querySelectorAll("tr")).filter(r => r.style.display !== "none");
  const dir  = (sortDir[col] = !sortDir[col]);
  rows.sort((a, b) => {{
    let av = a.cells[col].innerText.replace(/[▲▼週前%+,]/g,"").trim();
    let bv = b.cells[col].innerText.replace(/[▲▼週前%+,]/g,"").trim();
    let an = parseFloat(av), bn = parseFloat(bv);
    if (!isNaN(an) && !isNaN(bn)) return dir ? an - bn : bn - an;
    return dir ? av.localeCompare(bv,"zh") : bv.localeCompare(av,"zh");
  }});
  rows.forEach(r => tb.appendChild(r));
}}

// ── 搜尋 ──────────────────────────────────────────────────────────────────────
function filterTable() {{
  const q = document.getElementById("search").value.toLowerCase();
  document.querySelectorAll("#mainTable tbody tr").forEach(row => {{
    row.style.display = row.innerText.toLowerCase().includes(q) ? "" : "none";
  }});
}}

// ── 圖表 ──────────────────────────────────────────────────────────────────────
let mainChart = null, volChart = null;

function openChart(code, name) {{
  const data = CHART_DATA[code];
  if (!data) return;

  // 設定標題
  document.getElementById("modalTitle").innerHTML = `<span>${{code}}</span>${{name}}`;
  document.getElementById("yahooLink").href = `https://finance.yahoo.com/chart/${{code}}.TW`;
  document.getElementById("modalOverlay").classList.add("open");

  // 清除舊圖
  document.getElementById("main-chart").innerHTML = "";
  document.getElementById("vol-chart").innerHTML  = "";

  const commonOpts = {{
    layout:     {{ background: {{ color: "#0d1826" }}, textColor: "#94a3b8" }},
    grid:       {{ vertLines: {{ color: "#1a2d42" }}, horzLines: {{ color: "#1a2d42" }} }},
    timeScale:  {{ borderColor: "#2d3f55", timeVisible: true }},
    rightPriceScale: {{ borderColor: "#2d3f55" }},
    crosshair:  {{ mode: 1 }},
    handleScroll: true,
    handleScale:  true,
  }};

  // ── 主圖（K線 + BB + MA20）
  mainChart = LightweightCharts.createChart(
    document.getElementById("main-chart"),
    {{ ...commonOpts, height: 340 }}
  );

  const candleSeries = mainChart.addCandlestickSeries({{
    upColor:        "#f87171",
    downColor:      "#34d399",
    borderUpColor:  "#f87171",
    borderDownColor:"#34d399",
    wickUpColor:    "#f87171",
    wickDownColor:  "#34d399",
  }});
  candleSeries.setData(data.candles);

  // BB 填充區域（上下軌之間的帶狀）
  const bbUpperSeries = mainChart.addLineSeries({{
    color: "rgba(251,146,60,0.7)",
    lineWidth: 1,
    lineStyle: 2,       // dashed
    priceLineVisible: false,
    lastValueVisible: false,
    title: "BB上軌",
  }});
  bbUpperSeries.setData(data.bb_upper);

  const bbLowerSeries = mainChart.addLineSeries({{
    color: "rgba(167,139,250,0.7)",
    lineWidth: 1,
    lineStyle: 2,
    priceLineVisible: false,
    lastValueVisible: false,
    title: "BB下軌",
  }});
  bbLowerSeries.setData(data.bb_lower);

  const ma20Series = mainChart.addLineSeries({{
    color: "#3b82f6",
    lineWidth: 2,
    priceLineVisible: false,
    lastValueVisible: true,
    title: "MA20",
  }});
  ma20Series.setData(data.ma20);

  mainChart.timeScale().fitContent();

  // ── 成交量圖
  volChart = LightweightCharts.createChart(
    document.getElementById("vol-chart"),
    {{
      ...commonOpts,
      height: 100,
      rightPriceScale: {{ borderColor: "#2d3f55", scaleMargins: {{ top: 0.1, bottom: 0 }} }},
    }}
  );

  const volSeries = volChart.addHistogramSeries({{
    priceFormat: {{ type: "volume" }},
    priceScaleId: "right",
  }});
  volSeries.setData(data.volume);
  volChart.timeScale().fitContent();

  // 同步十字線
  mainChart.subscribeCrosshairMove(param => {{
    if (param.time) volChart.setCrosshairPosition(0, param.time, volSeries);
    else volChart.clearCrosshairPosition();
  }});
  volChart.subscribeCrosshairMove(param => {{
    if (param.time) mainChart.setCrosshairPosition(0, param.time, candleSeries);
    else mainChart.clearCrosshairPosition();
  }});
}}

function closeModal() {{
  document.getElementById("modalOverlay").classList.remove("open");
  document.querySelectorAll("#mainTable tbody tr").forEach(r => r.classList.remove("active"));
  if (mainChart) {{ mainChart.remove(); mainChart = null; }}
  if (volChart)  {{ volChart.remove();  volChart  = null; }}
}}

function closeModalOnBg(e) {{
  if (e.target === document.getElementById("modalOverlay")) closeModal();
}}

// 點擊列開圖
document.querySelectorAll("#mainTable tbody tr").forEach(row => {{
  row.addEventListener("click", () => {{
    document.querySelectorAll("#mainTable tbody tr").forEach(r => r.classList.remove("active"));
    row.classList.add("active");
    openChart(row.dataset.code, row.dataset.name);
  }});
}});

// ESC 關閉
document.addEventListener("keydown", e => {{ if (e.key === "Escape") closeModal(); }});
</script>
</body>
</html>
"""
    return html

# ── 10. 匯出 Excel ────────────────────────────────────────────────────────────
def save_excel(results, scan_time):
    """將掃描結果存成 Excel，檔名含日期，存於 scan.py 同目錄。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("   ⚠ openpyxl 未安裝，跳過 Excel 輸出（pip install openpyxl）", flush=True)
        return

    # ── 組成 DataFrame ──────────────────────────────────────────────────────
    rows = []
    for r in results:
        inst  = r.get("institutional") or {}
        tdcc  = r.get("tdcc")         or {}
        mc    = r.get("margin_chg")
        rv    = r.get("resist_vol",  0)
        sv    = r.get("support_vol", 0)
        vr    = r.get("vol_ratio")

        rows.append({
            "代碼":        r["code"],
            "名稱":        r["name"],
            "收盤":        r["close"],
            "MA20":        r["ma20"],
            "乖離%":       r["current_dev"],
            "峰乖離%":     r["peak_dev"],
            "距峰(週)":    r["weeks_since_peak"],
            "斜率%":       r["ma20_slope_pct"],
            "法人(萬股)":  inst.get("this_week"),
            "融資增減(張)": mc,
            "千張比%":     tdcc.get("pct"),
            "千張週變%":   tdcc.get("chg"),
            "壓力(萬張)":  round(rv / 1e7, 2) if rv else None,
            "支撐(萬張)":  round(sv / 1e7, 2) if sv else None,
            "支/壓":       vr,
        })

    df = pd.DataFrame(rows)

    # ── 寫入 Excel ──────────────────────────────────────────────────────────
    date_str  = datetime.now().strftime("%Y%m%d")
    base_dir  = os.path.dirname(os.path.abspath(__file__))
    xl_path   = os.path.join(base_dir, f"台股MA20掃描_{date_str}.xlsx")

    with pd.ExcelWriter(xl_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="MA20掃描結果", index=False, startrow=1)
        ws = writer.sheets["MA20掃描結果"]

        # ── 第 1 列：掃描時間說明 ───────────────────────────────────────────
        ws["A1"] = f"台股週K MA20 掃描結果　掃描時間：{scan_time}"
        ws["A1"].font      = Font(bold=True, color="FFFFFF", size=11)
        ws["A1"].fill      = PatternFill("solid", fgColor="1E3A5F")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
        ws.row_dimensions[1].height = 22

        # ── 第 2 列：欄位標題樣式 ───────────────────────────────────────────
        hdr_fill = PatternFill("solid", fgColor="0F2040")
        hdr_font = Font(bold=True, color="93C5FD", size=10)
        thin     = Side(style="thin", color="2D3F55")
        border   = Border(bottom=Side(style="medium", color="3B82F6"))

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 28

        # ── 資料列樣式 ─────────────────────────────────────────────────────
        fill_odd  = PatternFill("solid", fgColor="0D1826")
        fill_even = PatternFill("solid", fgColor="111F30")
        font_data = Font(color="E2E8F0", size=10)
        font_code = Font(color="FB923C", bold=True, size=10)

        # 數值欄索引（0-based）：乖離%以後都是數字
        num_cols = set(range(2, len(df.columns)))

        for row_idx in range(3, 3 + len(df)):
            fill = fill_odd if (row_idx % 2 == 1) else fill_even
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill      = fill
                cell.font      = font_code if col_idx == 1 else font_data
                cell.alignment = Alignment(
                    horizontal="right" if col_idx > 2 else "center",
                    vertical="center"
                )
                # 數值格式
                if col_idx in {3, 4}:                    # 收盤、MA20
                    cell.number_format = "#,##0.00"
                elif col_idx in {5, 6, 8, 11, 12}:      # 百分比欄
                    cell.number_format = "+0.00;-0.00;0.00"
                elif col_idx in {9, 10}:                 # 法人、融資
                    cell.number_format = "+#,##0.0;-#,##0.0;0"
                elif col_idx in {13, 14}:                # 壓力、支撐
                    cell.number_format = "#,##0.00"
                elif col_idx == 15:                      # 支/壓比值
                    cell.number_format = "0.00"

        # ── 凍結前兩欄（代碼、名稱） ───────────────────────────────────────
        ws.freeze_panes = "C3"

        # ── 自動篩選（排序＋篩選下拉） ─────────────────────────────────────
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A2:{last_col}2"

        # ── 自動調整欄寬 ───────────────────────────────────────────────────
        col_widths = {
            "代碼": 7, "名稱": 10, "收盤": 8, "MA20": 8,
            "乖離%": 7, "峰乖離%": 8, "距峰(週)": 8, "斜率%": 7,
            "法人(萬股)": 10, "融資增減(張)": 11,
            "千張比%": 8, "千張週變%": 8,
            "壓力(萬張)": 9, "支撐(萬張)": 9, "支/壓": 7,
        }
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 10)

    print(f"   Excel 已存：{xl_path}", flush=True)
    return xl_path

# ── 主程式 ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60, flush=True)
    print("  台股週K MA20 掃描工具 v4（大戶散戶週變化版）", flush=True)
    print("=" * 60, flush=True)

    # ① 取得上市股票清單
    stocks = get_twse_stock_list()
    if not stocks:
        print("無法取得股票清單，請確認網路連線後重試")
        sys.exit(1)

    # ② 掃描週K線篩選
    print(f"開始掃描（共 {len(stocks)} 檔）...", flush=True)
    results = scan_stocks(stocks, batch_size=150)
    print(f"初步篩選：{len(results)} 檔符合條件", flush=True)

    # ③ 取得三大法人週資料（大戶）
    print("取得大戶/散戶週變化資料...", flush=True)
    trading_days = get_trading_days(n=10)
    print(f"   交易日：{trading_days}", flush=True)
    institutional = fetch_institutional_weekly(trading_days)

    # ④ 取得融資週增減（散戶）
    margin_weekly = get_margin_weekly_change()

    # ⑤ 取得千張大戶比例（TDCC）
    tdcc_data = get_tdcc_weekly_change()

    # ⑥ 合併到 results
    for r in results:
        code = r["code"]
        r["institutional"] = institutional.get(code)
        r["margin_chg"]    = margin_weekly.get(code)
        r["tdcc"]          = tdcc_data.get(code)

    # ⑦ 產生 HTML
    scan_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    html = gen_html(results, scan_time)

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "index.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML 輸出：{out_path}", flush=True)

    # ⑧ 匯出 Excel（本機留存）
    print("匯出 Excel...", flush=True)
    save_excel(results, scan_time)

    print(f"完成！共 {len(results)} 檔。", flush=True)
